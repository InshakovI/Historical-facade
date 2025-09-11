import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                             QWidget, QLabel, QPushButton, QFileDialog, QListWidget,
                             QMessageBox, QComboBox, QSpinBox, QFrame)
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import Qt
import openpyxl
import qrcode
from PIL import Image
import io


class QRGeneratorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор QR-кодов из Excel")
        self.setGeometry(100, 100, 900, 700)
        
        # Основные переменные
        self.excel_file = None
        self.current_qr = None
        self.qr_codes = {}
        
        # Настройка стилей
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
                font-size: 14px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                padding: 8px 16px;
                font-size: 14px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QListWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            QComboBox {
                padding: 5px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            QComboBox:hover {
                background-color: #3498db;
                color: blue;
            }
            QComboBox QAbstractItemView {
                selection-background-color: #3498db;
                selection-color: blue;
            }
            QSpinBox {
                padding: 5px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            QFrame#footer {
                background-color: #e9e9e9;
                border-top: 1px solid #ddd;
                padding: 10px;
            }
        """)
        
        # Создаем основной виджет и layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # Заголовок
        header = QLabel("Генератор QR-кодов из Excel")
        header.setStyleSheet("""
            font-size: 20px;
            font-weight: bold;
            color: #2c3e50;
            padding: 10px;
        """)
        header.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(header)
        
        # 1. Секция загрузки файла
        file_frame = QFrame()
        file_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        file_layout = QHBoxLayout(file_frame)
        
        self.file_label = QLabel("Файл Excel не выбран")
        self.file_label.setStyleSheet("font-weight: bold; color: #3498db;")
        
        file_btn = QPushButton("Выбрать файл Excel")
        file_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        file_btn.clicked.connect(self.load_excel_file)
        
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_btn)
        main_layout.addWidget(file_frame)
        
        # 2. Секция настроек
        settings_frame = QFrame()
        settings_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        settings_layout = QHBoxLayout(settings_frame)
        
        # Выбор листа
        sheet_layout = QVBoxLayout()
        sheet_label = QLabel("Лист:")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setPlaceholderText("Выберите лист")
        self.sheet_combo.setStyleSheet("""
            QComboBox {
                min-width: 150px;
            }
        """)
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.sheet_combo)
        settings_layout.addLayout(sheet_layout)
        
        # Выбор колонки
        column_layout = QVBoxLayout()
        column_label = QLabel("Колонка с ФИО:")
        self.column_combo = QComboBox()
        self.column_combo.setPlaceholderText("Выберите колонку")
        self.column_combo.setStyleSheet("""
            QComboBox {
                min-width: 150px;
            }
        """)
        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_combo)
        settings_layout.addLayout(column_layout)
        
        # Размер QR-кода
        size_layout = QVBoxLayout()
        size_label = QLabel("Размер QR:")
        self.size_spin = QSpinBox()
        self.size_spin.setRange(100, 500)
        self.size_spin.setValue(250)
        size_layout.addWidget(size_label)
        size_layout.addWidget(self.size_spin)
        settings_layout.addLayout(size_layout)
        
        main_layout.addWidget(settings_frame)
        
        # Подключаем сигнал изменения листа
        self.sheet_combo.currentTextChanged.connect(self.update_columns)
        
        # 3. Секция генерации
        generate_btn = QPushButton("Сгенерировать QR-коды")
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                font-size: 16px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #d35400;
            }
        """)
        generate_btn.clicked.connect(self.generate_qr_codes)
        main_layout.addWidget(generate_btn, alignment=Qt.AlignCenter)
        
        # 4. Секция отображения данных
        data_frame = QFrame()
        data_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        data_layout = QHBoxLayout(data_frame)
        
        # Список ФИО
        self.data_list = QListWidget()
        self.data_list.setStyleSheet("font-size: 14px;")
        self.data_list.itemClicked.connect(self.show_qr_code)
        data_layout.addWidget(self.data_list, 1)
        
        # Область просмотра QR-кода
        self.qr_display = QLabel()
        self.qr_display.setAlignment(Qt.AlignCenter)
        self.qr_display.setMinimumSize(350, 350)
        self.qr_display.setStyleSheet("""
            background-color: white; 
            border: 2px solid #3498db;
            border-radius: 8px;
            padding: 10px;
        """)
        data_layout.addWidget(self.qr_display, 1)
        
        main_layout.addWidget(data_frame)
        
        # 5. Секция сохранения
        save_frame = QFrame()
        save_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        save_layout = QHBoxLayout(save_frame)
        
        save_current_btn = QPushButton("Сохранить текущий QR-код")
        save_current_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        save_current_btn.clicked.connect(self.save_current_qr)
        
        save_all_btn = QPushButton("Сохранить все QR-коды")
        save_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        save_all_btn.clicked.connect(self.save_all_qr)
        
        save_layout.addWidget(save_current_btn)
        save_layout.addWidget(save_all_btn)
        main_layout.addWidget(save_frame)
        
        # Футер с подписью по центру
        footer = QFrame()
        footer.setObjectName("footer")
        footer_layout = QHBoxLayout(footer)
        footer_label = QLabel("Разработка: Inshakov Igor")
        footer_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        footer_layout.addWidget(footer_label, alignment=Qt.AlignCenter)
        main_layout.addWidget(footer)
        
        # Устанавливаем отступы
        main_layout.setContentsMargins(20, 20, 20, 10)
        main_layout.setSpacing(15)
    
    def load_excel_file(self):
        """Загрузка Excel файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            
            # Загружаем доступные листы
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
    
    def update_columns(self):
        """Обновление списка колонок при выборе листа"""
        self.column_combo.clear()
        
        if not self.excel_file or not self.sheet_combo.currentText():
            return
            
        sheet_name = self.sheet_combo.currentText()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            # Получаем заголовки колонок из первой строки
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:  # Если ячейка не пустая
                    self.column_combo.addItem(str(cell.value), col_idx)
                else:
                    self.column_combo.addItem(f"Колонка {col_idx}", col_idx)
            
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать колонки: {str(e)}")
    
    def generate_qr_codes(self):
        """Генерация QR-кодов для выбранных данных"""
        if not all([self.excel_file, self.sheet_combo.currentText(), self.column_combo.currentText()]):
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл, лист и колонку")
            return
            
        self.data_list.clear()
        self.qr_codes = {}
        sheet_name = self.sheet_combo.currentText()
        column = self.column_combo.currentText()
        col_idx = self.column_combo.currentData()
        qr_size = self.size_spin.value()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            # Читаем данные из колонки (пропускаем заголовок)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= col_idx:
                    value = row[col_idx-1]
                    if value is not None:
                        value = str(value).strip()
                        if value:
                            # Генерируем QR-код
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_H,
                                box_size=12,
                                border=4,
                            )
                            qr.add_data(value)
                            qr.make(fit=True)
                            
                            img = qr.make_image(fill_color="black", back_color="white")
                            img = img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
                            
                            # Сохраняем в словарь
                            self.qr_codes[value] = img
                            self.data_list.addItem(value)
            
            wb.close()
            QMessageBox.information(self, "Готово", f"Сгенерировано {len(self.qr_codes)} QR-кодов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать QR-коды: {str(e)}")
    
    def show_qr_code(self, item):
        """Отображение выбранного QR-кода"""
        text = item.text()
        if text in self.qr_codes:
            self.current_qr = self.qr_codes[text]
            
            # Конвертируем PIL.Image в QPixmap через байты
            img_byte_arr = io.BytesIO()
            self.current_qr.save(img_byte_arr, format='PNG')
            qimage = QImage.fromData(img_byte_arr.getvalue())
            pixmap = QPixmap.fromImage(qimage)
            
            # Масштабируем с сохранением пропорций
            display_size = min(self.qr_display.width(), self.qr_display.height()) - 20
            pixmap = pixmap.scaled(
                display_size, 
                display_size,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            self.qr_display.setPixmap(pixmap)
    
    def save_current_qr(self):
        """Сохранение текущего QR-кода"""
        if not hasattr(self, 'current_qr') or not self.current_qr:
            QMessageBox.warning(self, "Ошибка", "Не выбран QR-код для сохранения")
            return
            
        current_item = self.data_list.currentItem()
        if not current_item:
            return
            
        text = current_item.text()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить QR-код", f"{text}.png", "PNG Files (*.png)"
        )
        
        if file_path:
            try:
                self.current_qr.save(file_path)
                QMessageBox.information(self, "Успешно", "QR-код сохранен")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
    
    def save_all_qr(self):
        """Сохранение всех QR-кодов в выбранную папку"""
        if not self.qr_codes:
            QMessageBox.warning(self, "Ошибка", "Нет QR-кодов для сохранения")
            return
            
        dir_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        
        if dir_path:
            try:
                saved_count = 0
                for text, img in self.qr_codes.items():
                    safe_text = "".join(c for c in text if c.isalnum() or c in (' ', '_')).rstrip()
                    file_path = os.path.join(dir_path, f"{safe_text}.png")
                    img.save(file_path, "PNG")
                    saved_count += 1
                
                QMessageBox.information(self, "Успешно", f"Сохранено {saved_count} QR-кодов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файлы: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = QRGeneratorApp()
    window.show()
    sys.exit(app.exec_())
    sys.exit(app.exec_())