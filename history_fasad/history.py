import sys
import os
import sqlite3
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                            QLabel, QPushButton, QWidget, QTableWidget, 
                            QTableWidgetItem, QMessageBox, QLineEdit, 
                            QStackedWidget, QListWidget, QComboBox, 
                            QSpinBox, QFrame, QFileDialog, QGroupBox)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon, QPixmap, QFont, QImage, QPalette, QColor
import qrcode
from pyzbar.pyzbar import decode
import cv2
import openpyxl
from PIL import Image
import io
import pandas as pd
import uuid

class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("QR Code Manager")
        self.setGeometry(100, 100, 1000, 750)
        
        # Генерация уникального имени базы данных при каждом запуске
        self.db_name = f"visitors_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        
        # Центральный виджет и макет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # Заголовок
        header = QLabel("QR Code Manager - Учет посетителей и генерация QR-кодов")
        header.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 15px;
                text-align: center;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3498db, stop:1 #2c3e50);
                border-radius: 10px;
                color: white;
            }
        """)
        header.setMinimumHeight(60)
        main_layout.addWidget(header)
        
        # Кнопки выбора режима
        mode_frame = QFrame()
        mode_frame.setStyleSheet("""
            QFrame {
                background-color: #ecf0f1;
                border-radius: 8px;
                padding: 5px;
            }
        """)
        mode_layout = QHBoxLayout(mode_frame)
        mode_layout.setContentsMargins(10, 10, 10, 10)
        mode_layout.setSpacing(20)
        
        self.scanner_btn = QPushButton("📷 Режим сканирования")
        self.scanner_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 15px 25px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        
        self.generator_btn = QPushButton("🎨 Режим генерации")
        self.generator_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 15px 25px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
            QPushButton:pressed {
                background-color: #219653;
            }
        """)
        
        mode_layout.addWidget(self.scanner_btn)
        mode_layout.addWidget(self.generator_btn)
        main_layout.addWidget(mode_frame)
        
        # Стек виджетов для переключения между режимами
        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget, 1)
        
        # Инициализация обоих режимов
        self.init_scanner_mode()
        self.init_generator_mode()
        
        # Подключение кнопок
        self.scanner_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))
        self.generator_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))
        
        # Футер
        footer = QFrame()
        footer.setObjectName("footer")
        footer.setStyleSheet("""
            QFrame#footer {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2c3e50, stop:1 #34495e);
                border-radius: 8px;
                padding: 12px;
            }
        """)
        footer_layout = QHBoxLayout(footer)
        footer_label = QLabel(f"Разработка: Inshakov Igor | Версия 1.1 | База: {self.db_name}")
        footer_label.setStyleSheet("color: #bdc3c7; font-size: 12px; font-weight: bold;")
        footer_layout.addWidget(footer_label, alignment=Qt.AlignCenter)
        main_layout.addWidget(footer)
        
        # Установка общего стиля приложения
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8f9fa;
            }
            QWidget {
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QLabel {
                color: #2c3e50;
                font-size: 14px;
            }
            QLineEdit, QComboBox, QSpinBox {
                padding: 8px;
                font-size: 14px;
                border: 2px solid #bdc3c7;
                border-radius: 6px;
                background-color: white;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
                border-color: #3498db;
            }
            QTableWidget {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                font-size: 14px;
                background-color: white;
                gridline-color: #ecf0f1;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QListWidget {
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                background-color: white;
                font-size: 14px;
            }
            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid #ecf0f1;
            }
            QListWidget::item:selected {
                background-color: #3498db;
                color: white;
                border-radius: 4px;
            }
        """)
    
    def init_scanner_mode(self):
        """Инициализация режима сканирования"""
        self.scanner_widget = QWidget()
        scanner_layout = QHBoxLayout(self.scanner_widget)
        scanner_layout.setSpacing(15)
        scanner_layout.setContentsMargins(10, 10, 10, 10)
        
        # Левая панель (сканер)
        left_panel = QGroupBox("Сканер QR-кодов")
        left_panel.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #3498db;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        left_layout = QVBoxLayout(left_panel)
        
        # Правая панель (информация)
        right_panel = QGroupBox("Информация о посетителях")
        right_panel.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #2ecc71;
                border-radius: 10px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        right_layout = QVBoxLayout(right_panel)
        
        # Добавляем панели в главный макет
        scanner_layout.addWidget(left_panel, 1)
        scanner_layout.addWidget(right_panel, 2)
        
        # Настройка левой панели
        # Выбор камеры
        camera_select_layout = QHBoxLayout()
        camera_label = QLabel("Выберите камеру:")
        self.camera_combo = QComboBox()
        self.camera_combo.setMinimumWidth(200)
        self.refresh_cameras_btn = QPushButton("🔄")
        self.refresh_cameras_btn.setFixedSize(30, 30)
        self.refresh_cameras_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        
        camera_select_layout.addWidget(camera_label)
        camera_select_layout.addWidget(self.camera_combo)
        camera_select_layout.addWidget(self.refresh_cameras_btn)
        left_layout.addLayout(camera_select_layout)
        
        self.camera_label = QLabel()
        self.camera_label.setAlignment(Qt.AlignCenter)
        self.camera_label.setMinimumSize(400, 300)
        self.camera_label.setStyleSheet("""
            QLabel {
                background-color: #2c3e50;
                border: 3px solid #34495e;
                border-radius: 8px;
                color: white;
                font-weight: bold;
            }
        """)
        self.camera_label.setText("Камера не активирована")
        
        self.scan_button = QPushButton("▶ Начать сканирование")
        self.scan_button.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                border: none;
                color: white;
                padding: 12px 24px;
                text-align: center;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #219653;
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        
        left_layout.addWidget(self.camera_label)
        left_layout.addWidget(self.scan_button)
        left_layout.addStretch()
        
        # Настройка правой панели
        # Поле для ввода мероприятия
        event_layout = QHBoxLayout()
        self.event_label = QLabel("Название мероприятия:")
        self.event_input = QLineEdit()
        self.event_input.setPlaceholderText("Введите название мероприятия...")
        event_layout.addWidget(self.event_label)
        event_layout.addWidget(self.event_input)
        right_layout.addLayout(event_layout)
        
        self.visitor_table = QTableWidget()
        self.visitor_table.setColumnCount(4)
        self.visitor_table.setHorizontalHeaderLabels(["ФИО", "Организация", "Время посещения", "Мероприятие"])
        self.visitor_table.horizontalHeader().setStretchLastSection(True)
        
        button_layout = QHBoxLayout()
        self.export_button = QPushButton("📊 Экспорт в Excel")
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                border: none;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        self.clear_db_button = QPushButton("🗑️ Очистить базу")
        self.clear_db_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                border: none;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.clear_db_button)
        button_layout.addStretch()
        
        right_layout.addWidget(self.visitor_table)
        right_layout.addLayout(button_layout)
        
        # Инициализация базы данных
        self.init_db()
        
        # Настройка камеры
        self.camera = None
        self.timer = QTimer()
        self.scanning = False
        
        # Заполняем список камер
        self.refresh_cameras()
        
        # Подключение сигналов
        self.scan_button.clicked.connect(self.toggle_scan)
        self.timer.timeout.connect(self.update_frame)
        self.export_button.clicked.connect(self.export_to_excel)
        self.clear_db_button.clicked.connect(self.clear_database)
        self.refresh_cameras_btn.clicked.connect(self.refresh_cameras)
        
        # Добавляем виджет в стек
        self.stacked_widget.addWidget(self.scanner_widget)
    
    def refresh_cameras(self):
        """Обновление списка доступных камер"""
        self.camera_combo.clear()
        
        # Проверяем доступные камеры
        index = 0
        cameras = []
        while True:
            cap = cv2.VideoCapture(index)
            if not cap.read()[0]:
                break
            else:
                cameras.append(index)
            cap.release()
            index += 1
        
        if cameras:
            for cam_index in cameras:
                self.camera_combo.addItem(f"Камера {cam_index}", cam_index)
        else:
            self.camera_combo.addItem("Камеры не найдены", -1)
    
    def init_generator_mode(self):
        """Инициализация режима генерации"""
        self.generator_widget = QWidget()
        generator_layout = QVBoxLayout(self.generator_widget)
        generator_layout.setSpacing(15)
        generator_layout.setContentsMargins(10, 10, 10, 10)
        
        # 1. Секция загрузки файла
        file_frame = QGroupBox("Загрузка данных")
        file_frame.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #f39c12;
                border-radius: 8px;
                padding-top: 15px;
            }
        """)
        file_layout = QHBoxLayout(file_frame)
        
        self.file_label = QLabel("Файл Excel не выбран")
        self.file_label.setStyleSheet("font-weight: bold; color: #e67e22;")
        
        file_btn = QPushButton("📁 Выбрать файл Excel")
        file_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                border: none;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
        """)
        
        file_layout.addWidget(self.file_label, 1)
        file_layout.addWidget(file_btn)
        generator_layout.addWidget(file_frame)
        
        # 2. Секция настроек
        settings_frame = QGroupBox("Настройки генерации")
        settings_frame.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #9b59b6;
                border-radius: 8px;
                padding-top: 15px;
            }
        """)
        settings_layout = QHBoxLayout(settings_frame)
        
        # Выбор листа
        sheet_layout = QVBoxLayout()
        sheet_label = QLabel("Лист:")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setPlaceholderText("Выберите лист")
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.sheet_combo)
        settings_layout.addLayout(sheet_layout)
        
        # Выбор колонки
        column_layout = QVBoxLayout()
        column_label = QLabel("Колонка с ФИО:")
        self.column_combo = QComboBox()
        self.column_combo.setPlaceholderText("Выберите колонку")
        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_combo)
        settings_layout.addLayout(column_layout)
        
        # Размер QR-кода
        size_layout = QVBoxLayout()
        size_label = QLabel("Размер QR:")
        self.size_spin = QSpinBox()
        self.size_spin.setRange(100, 500)
        self.size_spin.setValue(250)
        self.size_spin.setSuffix(" px")
        size_layout.addWidget(size_label)
        size_layout.addWidget(self.size_spin)
        settings_layout.addLayout(size_layout)
        
        generator_layout.addWidget(settings_frame)
        
        # 3. Секция генерации
        generate_btn = QPushButton("⚡ Сгенерировать QR-коды")
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                border: none;
                color: white;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
        """)
        generator_layout.addWidget(generate_btn, alignment=Qt.AlignCenter)
        
        # 4. Секция отображения данных
        data_frame = QGroupBox("Данные и предпросмотр")
        data_frame.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #2c3e50;
                border-radius: 8px;
                padding-top: 15px;
            }
        """)
        data_layout = QHBoxLayout(data_frame)
        
        # Список ФИО
        list_frame = QVBoxLayout()
        list_label = QLabel("Список записей:")
        self.data_list = QListWidget()
        list_frame.addWidget(list_label)
        list_frame.addWidget(self.data_list)
        data_layout.addLayout(list_frame, 1)
        
        # Область просмотра QR-кода
        qr_frame = QVBoxLayout()
        qr_label = QLabel("Предпросмотр QR-кода:")
        self.qr_display = QLabel()
        self.qr_display.setAlignment(Qt.AlignCenter)
        self.qr_display.setMinimumSize(300, 300)
        self.qr_display.setStyleSheet("""
            QLabel {
                background-color: white;
                border: 3px solid #bdc3c7;
                border-radius: 10px;
                padding: 15px;
            }
        """)
        self.qr_display.setText("QR-код не выбран")
        qr_frame.addWidget(qr_label)
        qr_frame.addWidget(self.qr_display)
        data_layout.addLayout(qr_frame, 1)
        generator_layout.addWidget(data_frame)
        
        # 5. Секция сохранения
        save_frame = QGroupBox("Сохранение QR-кодов")
        save_frame.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
                border: 2px solid #27ae60;
                border-radius: 8px;
                padding-top: 15px;
            }
        """)
        save_layout = QHBoxLayout(save_frame)
        
        save_current_btn = QPushButton("💾 Сохранить текущий QR-код")
        save_current_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                border: none;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        save_all_btn = QPushButton("📦 Сохранить все QR-коды")
        save_all_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                border: none;
                color: white;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        
        save_layout.addWidget(save_current_btn)
        save_layout.addWidget(save_all_btn)
        generator_layout.addWidget(save_frame)
        
        # Основные переменные
        self.excel_file = None
        self.current_qr = None
        self.qr_codes = {}
        
        # Подключение сигналов
        file_btn.clicked.connect(self.load_excel_file)
        self.sheet_combo.currentTextChanged.connect(self.update_columns)
        generate_btn.clicked.connect(self.generate_qr_codes)
        self.data_list.itemClicked.connect(self.show_qr_code)
        save_current_btn.clicked.connect(self.save_current_qr)
        save_all_btn.clicked.connect(self.save_all_qr)
        
        # Добавляем виджет в стек
        self.stacked_widget.addWidget(self.generator_widget)
    
    # Методы для режима сканирования
    def init_db(self):
        self.conn = sqlite3.connect(self.db_name)
        self.cursor = self.conn.cursor()
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS visitors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT,
                organization TEXT,
                qr_data TEXT UNIQUE,
                visit_time DATETIME,
                event TEXT
            )
        ''')
        self.conn.commit()
        self.update_visitors_table()
    
    def clear_database(self):
        """Очистка базы данных"""
        reply = QMessageBox.question(
            self, "Подтверждение", 
            "Вы уверены, что хотите очистить базу данных? Все данные будут удалены.",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.cursor.execute("DELETE FROM visitors")
                self.conn.commit()
                self.update_visitors_table()
                QMessageBox.information(self, "Успешно", "База данных очищена")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось очистить базу: {str(e)}")
    
    def toggle_scan(self):
        if not self.scanning:
            # Получаем выбранную камеру
            camera_index = self.camera_combo.currentData()
            if camera_index == -1:
                QMessageBox.warning(self, "Ошибка", "Не найдено доступных камер!")
                return
            
            try:
                self.camera = cv2.VideoCapture(camera_index)
                if not self.camera.isOpened():
                    QMessageBox.warning(self, "Ошибка", "Не удалось открыть камеру!")
                    return
                
                self.scanning = True
                self.scan_button.setText("⏹️ Остановить сканирование")
                self.scan_button.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        border: none;
                        color: white;
                        padding: 12px 24px;
                        text-align: center;
                        font-size: 16px;
                        font-weight: bold;
                        border-radius: 8px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                self.timer.start(20)
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при запуске камеры: {str(e)}")
        else:
            self.scanning = False
            self.scan_button.setText("▶ Начать сканирование")
            self.scan_button.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    border: none;
                    color: white;
                    padding: 12px 24px;
                    text-align: center;
                    font-size: 16px;
                    font-weight: bold;
                    border-radius: 8px;
                }
                QPushButton:hover {
                    background-color: #219653;
                }
            """)
            self.timer.stop()
            if self.camera and self.camera.isOpened():
                self.camera.release()
            self.camera_label.clear()
            self.camera_label.setText("Камера отключена")
    
    def update_frame(self):
        if self.camera and self.camera.isOpened():
            ret, frame = self.camera.read()
            if ret:
                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                decoded_objects = decode(frame)
                
                if decoded_objects:
                    qr_data = decoded_objects[0].data.decode('utf-8')
                    self.process_qr_code(qr_data)
                
                h, w, ch = frame_rgb.shape
                bytes_per_line = ch * w
                q_img = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.camera_label.setPixmap(QPixmap.fromImage(q_img).scaled(
                    self.camera_label.width(), 
                    self.camera_label.height(), 
                    Qt.KeepAspectRatio
                ))
    
    def process_qr_code(self, qr_data):
        try:
            self.cursor.execute("SELECT * FROM visitors WHERE qr_data=?", (qr_data,))
            visitor = self.cursor.fetchone()
            
            if visitor:
                visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                event = self.event_input.text() or "Не указано"
                self.cursor.execute(
                    "UPDATE visitors SET visit_time=?, event=? WHERE qr_data=?",
                    (visit_time, event, qr_data)
                )
                self.conn.commit()
                
                QMessageBox.information(
                    self, "Успешно", 
                    f"Посетитель {visitor[1]} уже зарегистрирован.\nВремя обновлено."
                )
            else:
                visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                event = self.event_input.text() or "Не указано"
                
                parts = qr_data.split(';')
                full_name = parts[0] if len(parts) > 0 else "Неизвестный"
                organization = parts[1] if len(parts) > 1 else "Не указана"
                
                self.cursor.execute(
                    "INSERT INTO visitors (full_name, organization, qr_data, visit_time, event) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (full_name, organization, qr_data, visit_time, event)
                )
                self.conn.commit()
                
                QMessageBox.information(
                    self, "Успешно", 
                    f"Новый посетитель {full_name} зарегистрирован!"
                )
            
            self.update_visitors_table()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
    
    def update_visitors_table(self):
        self.cursor.execute("SELECT full_name, organization, visit_time, event FROM visitors ORDER BY visit_time DESC")
        visitors = self.cursor.fetchall()
        
        self.visitor_table.setRowCount(len(visitors))
        for row, visitor in enumerate(visitors):
            for col, data in enumerate(visitor):
                item = QTableWidgetItem(str(data))
                item.setTextAlignment(Qt.AlignCenter)
                self.visitor_table.setItem(row, col, item)
        
        self.visitor_table.resizeColumnsToContents()
    
    def export_to_excel(self):
        try:
            self.cursor.execute("SELECT full_name, organization, visit_time, event FROM visitors")
            visitors = self.cursor.fetchall()
            
            if not visitors:
                QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта")
                return
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Сохранить как Excel", "visitors.xlsx", "Excel Files (*.xlsx)"
            )
            
            if file_path:
                df = pd.DataFrame(visitors, columns=["ФИО", "Организация", "Время посещения", "Мероприятие"])
                df.to_excel(file_path, index=False)
                
                QMessageBox.information(self, "Успешно", f"Данные экспортированы в {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")
    
    # Методы для режима генерации
    def load_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.file_label.setStyleSheet("font-weight: bold; color: #27ae60;")
            
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
    
    def update_columns(self):
        self.column_combo.clear()
        
        if not self.excel_file or not self.sheet_combo.currentText():
            return
            
        sheet_name = self.sheet_combo.currentText()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    self.column_combo.addItem(str(cell.value), col_idx)
                else:
                    self.column_combo.addItem(f"Колонка {col_idx}", col_idx)
            
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать колонки: {str(e)}")
    
    def generate_qr_codes(self):
        if not all([self.excel_file, self.sheet_combo.currentText(), self.column_combo.currentText()]):
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл, лист и колонку")
            return
            
        self.data_list.clear()
        self.qr_codes = {}
        sheet_name = self.sheet_combo.currentText()
        column = self.column_combo.currentText()
        col_idx = self.column_combo.currentData()
        qr_size = self.size_spin.value()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= col_idx:
                    value = row[col_idx-1]
                    if value is not None:
                        value = str(value).strip()
                        if value:
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_H,
                                box_size=12,
                                border=4,
                            )
                            qr.add_data(value)
                            qr.make(fit=True)
                            
                            img = qr.make_image(fill_color="black", back_color="white")
                            img = img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
                            
                            self.qr_codes[value] = img
                            self.data_list.addItem(value)
            
            wb.close()
            QMessageBox.information(self, "Готово", f"Сгенерировано {len(self.qr_codes)} QR-кодов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать QR-коды: {str(e)}")
    
    def show_qr_code(self, item):
        text = item.text()
        if text in self.qr_codes:
            self.current_qr = self.qr_codes[text]
            
            img_byte_arr = io.BytesIO()
            self.current_qr.save(img_byte_arr, format='PNG')
            qimage = QImage.fromData(img_byte_arr.getvalue())
            pixmap = QPixmap.fromImage(qimage)
            
            display_size = min(self.qr_display.width(), self.qr_display.height()) - 20
            pixmap = pixmap.scaled(
                display_size, 
                display_size,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            self.qr_display.setPixmap(pixmap)
    
    def save_current_qr(self):
        if not hasattr(self, 'current_qr') or not self.current_qr:
            QMessageBox.warning(self, "Ошибка", "Не выбран QR-код для сохранения")
            return
            
        current_item = self.data_list.currentItem()
        if not current_item:
            return
            
        text = current_item.text()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить QR-код", f"{text}.png", "PNG Files (*.png)"
        )
        
        if file_path:
            try:
                self.current_qr.save(file_path)
                QMessageBox.information(self, "Успешно", "QR-код сохранен")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
    
    def save_all_qr(self):
        if not self.qr_codes:
            QMessageBox.warning(self, "Ошибка", "Нет QR-кодов для сохранения")
            return
            
        dir_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        
        if dir_path:
            try:
                saved_count = 0
                for text, img in self.qr_codes.items():
                    safe_text = "".join(c for c in text if c.isalnum() or c in (' ', '_')).rstrip()
                    file_path = os.path.join(dir_path, f"{safe_text}.png")
                    img.save(file_path)
                    saved_count += 1
                
                QMessageBox.information(self, "Успешно", f"Сохранено {saved_count} QR-кодов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файлы: {str(e)}")
    
    def closeEvent(self, event):
        if self.camera and self.camera.isOpened():
            self.camera.release()
        if hasattr(self, 'conn'):
            self.conn.close()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())