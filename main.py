import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                             QWidget, QLabel, QPushButton, QFileDialog, QListWidget,
                             QMessageBox, QComboBox, QSpinBox)
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtCore import Qt, QByteArray, QBuffer
import openpyxl
import qrcode
from PIL import Image


class QRGeneratorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Генератор QR-кодов из Excel")
        self.setGeometry(100, 100, 800, 600)
        
        # Основные переменные
        self.excel_file = None
        self.sheet_name = None
        self.column_name = None
        self.qr_codes = {}
        
        # Создаем основной виджет и layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 1. Секция загрузки файла
        file_layout = QHBoxLayout()
        self.file_label = QLabel("Файл Excel не выбран")
        file_btn = QPushButton("Выбрать файл Excel")
        file_btn.clicked.connect(self.load_excel_file)
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_btn)
        main_layout.addLayout(file_layout)
        
        # 2. Секция настроек
        settings_layout = QHBoxLayout()
        
        # Выбор листа
        self.sheet_combo = QComboBox()
        self.sheet_combo.setPlaceholderText("Выберите лист")
        sheet_label = QLabel("Лист:")
        settings_layout.addWidget(sheet_label)
        settings_layout.addWidget(self.sheet_combo)
        
        # Выбор колонки
        self.column_combo = QComboBox()
        self.column_combo.setPlaceholderText("Выберите колонку")
        column_label = QLabel("Колонка с ФИО:")
        settings_layout.addWidget(column_label)
        settings_layout.addWidget(self.column_combo)
        
        # Размер QR-кода
        size_label = QLabel("Размер QR:")
        self.size_spin = QSpinBox()
        self.size_spin.setRange(100, 500)
        self.size_spin.setValue(200)
        settings_layout.addWidget(size_label)
        settings_layout.addWidget(self.size_spin)
        
        main_layout.addLayout(settings_layout)
        
        # Подключаем сигнал изменения листа
        self.sheet_combo.currentTextChanged.connect(self.update_columns)
        
        # 3. Секция генерации
        generate_btn = QPushButton("Сгенерировать QR-коды")
        generate_btn.clicked.connect(self.generate_qr_codes)
        main_layout.addWidget(generate_btn)
        
        # 4. Секция отображения данных
        data_layout = QHBoxLayout()
        
        # Список ФИО
        self.data_list = QListWidget()
        self.data_list.itemClicked.connect(self.show_qr_code)
        data_layout.addWidget(self.data_list, 1)
        
        # Область просмотра QR-кода
        self.qr_display = QLabel()
        self.qr_display.setAlignment(Qt.AlignCenter)
        self.qr_display.setMinimumSize(300, 300)
        data_layout.addWidget(self.qr_display, 1)
        
        main_layout.addLayout(data_layout)
        
        # 5. Секция сохранения
        save_layout = QHBoxLayout()
        save_current_btn = QPushButton("Сохранить текущий QR-код")
        save_current_btn.clicked.connect(self.save_current_qr)
        save_all_btn = QPushButton("Сохранить все QR-коды")
        save_all_btn.clicked.connect(self.save_all_qr)
        save_layout.addWidget(save_current_btn)
        save_layout.addWidget(save_all_btn)
        main_layout.addLayout(save_layout)
    
    def load_excel_file(self):
        """Загрузка Excel файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            
            # Загружаем доступные листы
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
    
    def update_columns(self):
        """Обновление списка колонок при выборе листа"""
        self.column_combo.clear()
        
        if not self.excel_file or not self.sheet_combo.currentText():
            return
            
        sheet_name = self.sheet_combo.currentText()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            # Получаем заголовки колонок из первой строки
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:  # Если ячейка не пустая
                    self.column_combo.addItem(str(cell.value), col_idx)
                else:
                    self.column_combo.addItem(f"Колонка {col_idx}", col_idx)
            
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать колонки: {str(e)}")
    
    def generate_qr_codes(self):
        """Генерация QR-кодов для выбранных данных"""
        if not all([self.excel_file, self.sheet_combo.currentText(), self.column_combo.currentText()]):
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл, лист и колонку")
            return
            
        self.data_list.clear()
        self.qr_codes = {}
        sheet_name = self.sheet_combo.currentText()
        column = self.column_combo.currentText()
        col_idx = self.column_combo.currentData()  # Получаем индекс колонки
        qr_size = self.size_spin.value()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            # Читаем данные из колонки (пропускаем заголовок)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= col_idx:  # Проверяем, что строка имеет достаточно колонок
                    value = row[col_idx-1]
                    if value is not None:
                        value = str(value).strip()
                        if value:  # Если значение не пустое
                            # Генерируем QR-код
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_L,
                                box_size=10,
                                border=4,
                            )
                            qr.add_data(value)
                            qr.make(fit=True)
                            
                            img = qr.make_image(fill_color="black", back_color="white")
                            img = img.resize((qr_size, qr_size))
                            
                            # Сохраняем в словарь
                            self.qr_codes[value] = img
                            self.data_list.addItem(value)
            
            wb.close()
            QMessageBox.information(self, "Готово", f"Сгенерировано {len(self.qr_codes)} QR-кодов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать QR-коды: {str(e)}")
    
    def show_qr_code(self, item):
        """Отображение выбранного QR-кода"""
        text = item.text()
        if text in self.qr_codes:
            img = self.qr_codes[text]
            
            # Конвертируем PIL.Image в QPixmap
            qimage = QImage(img.tobytes(), img.size[0], img.size[1], QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(qimage)
            
            # Масштабируем изображение для отображения
            if not pixmap.isNull():
                pixmap = pixmap.scaled(
                    self.qr_display.width(), 
                    self.qr_display.height(),
                    Qt.KeepAspectRatio
                )
                self.qr_display.setPixmap(pixmap)
            else:
                self.qr_display.clear()
    
    def save_current_qr(self):
        """Сохранение текущего QR-кода"""
        current_item = self.data_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "Ошибка", "Не выбран QR-код для сохранения")
            return
            
        text = current_item.text()
        if text not in self.qr_codes:
            return
            
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить QR-код", f"{text}.png", "PNG Files (*.png)"
        )
        
        if file_path:
            try:
                self.qr_codes[text].save(file_path)
                QMessageBox.information(self, "Успешно", "QR-код сохранен")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
    
    def save_all_qr(self):
        """Сохранение всех QR-кодов в выбранную папку"""
        if not self.qr_codes:
            QMessageBox.warning(self, "Ошибка", "Нет QR-кодов для сохранения")
            return
            
        dir_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        
        if dir_path:
            try:
                saved_count = 0
                for text, img in self.qr_codes.items():
                    # Очищаем имя файла от недопустимых символов
                    safe_text = "".join(c for c in text if c.isalnum() or c in (' ', '_')).rstrip()
                    file_path = os.path.join(dir_path, f"{safe_text}.png")
                    img.save(file_path)
                    saved_count += 1
                
                QMessageBox.information(self, "Успешно", f"Сохранено {saved_count} QR-кодов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файлы: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = QRGeneratorApp()
    window.show()
    sys.exit(app.exec_())