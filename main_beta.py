import sys
import os
import sqlite3
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, 
                            QLabel, QPushButton, QWidget, QTableWidget, 
                            QTableWidgetItem, QMessageBox, QLineEdit, 
                            QStackedWidget, QListWidget, QComboBox, 
                            QSpinBox, QFrame, QFileDialog)
from PyQt5.QtCore import Qt, QTimer
from PyQt5.QtGui import QIcon, QPixmap, QFont, QImage
import qrcode
from pyzbar.pyzbar import decode
import cv2
import openpyxl
from PIL import Image
import io
import pandas as pd

class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("QR Code Manager")
        self.setGeometry(100, 100, 950, 700)
        
        # Центральный виджет и макет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Заголовок
        header = QLabel("QR Code Manager - Учет посетителей и генерация QR-кодов")
        header.setStyleSheet("""
            font-size: 22px;
            font-weight: bold;
            color: #2c3e50;
            padding: 15px;
            text-align: center;
        """)
        main_layout.addWidget(header)
        
        # Кнопки выбора режима
        mode_frame = QFrame()
        mode_layout = QHBoxLayout(mode_frame)
        mode_layout.setContentsMargins(0, 0, 0, 0)
        
        self.scanner_btn = QPushButton("Режим сканирования")
        self.scanner_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        
        self.generator_btn = QPushButton("Режим генерации")
        self.generator_btn.setStyleSheet("""
            QPushButton {
                background-color: #2ecc71;
                color: white;
                border: none;
                padding: 12px 24px;
                font-size: 16px;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #27ae60;
            }
        """)
        
        mode_layout.addWidget(self.scanner_btn)
        mode_layout.addWidget(self.generator_btn)
        main_layout.addWidget(mode_frame)
        
        # Стек виджетов для переключения между режимами
        self.stacked_widget = QStackedWidget()
        main_layout.addWidget(self.stacked_widget, 1)
        
        # Инициализация обоих режимов
        self.init_scanner_mode()
        self.init_generator_mode()
        
        # Подключение кнопок
        self.scanner_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(0))
        self.generator_btn.clicked.connect(lambda: self.stacked_widget.setCurrentIndex(1))
        
        # Футер
        footer = QFrame()
        footer.setObjectName("footer")
        footer.setStyleSheet("""
            QFrame#footer {
                background-color: #e9e9e9;
                border-top: 1px solid #ddd;
                padding: 10px;
            }
        """)
        footer_layout = QHBoxLayout(footer)
        footer_label = QLabel("Разработка: Inshakov Igor | Версия 1.0")
        footer_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        footer_layout.addWidget(footer_label, alignment=Qt.AlignCenter)
        main_layout.addWidget(footer)
        
        # Установка стиля приложения
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
                font-size: 14px;
            }
            QTableWidget {
                border: 1px solid #ddd;
                font-size: 14px;
            }
            QHeaderView::section {
                background-color: #f8f8f8;
                padding: 5px;
                border: 1px solid #ddd;
            }
            QLineEdit {
                padding: 5px;
                font-size: 14px;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
        """)
    
    def init_scanner_mode(self):
        """Инициализация режима сканирования"""
        self.scanner_widget = QWidget()
        scanner_layout = QHBoxLayout(self.scanner_widget)
        
        # Левая панель (сканер)
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        
        # Правая панель (информация)
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        # Добавляем панели в главный макет
        scanner_layout.addWidget(left_panel, 1)
        scanner_layout.addWidget(right_panel, 2)
        
        # Настройка левой панели
        self.camera_label = QLabel()
        self.camera_label.setAlignment(Qt.AlignCenter)
        self.camera_label.setMinimumSize(400, 300)
        self.camera_label.setStyleSheet("background-color: #f0f0f0; border: 2px solid #ccc;")
        
        self.scan_button = QPushButton("Начать сканирование")
        self.scan_button.setStyleSheet(
            "QPushButton {"
            "background-color: #4CAF50;"
            "border: none;"
            "color: white;"
            "padding: 10px 24px;"
            "text-align: center;"
            "font-size: 16px;"
            "border-radius: 4px;"
            "}"
            "QPushButton:hover {"
            "background-color: #45a049;"
            "}"
        )
        
        left_layout.addWidget(QLabel("<h2 style='color: #333;'>Сканер QR-кодов</h2>"))
        left_layout.addWidget(self.camera_label)
        left_layout.addWidget(self.scan_button)
        left_layout.addStretch()
        
        # Настройка правой панели
        self.info_label = QLabel("<h3 style='color: #333;'>Информация о посетителе</h3>")
        
        # Поле для ввода мероприятия
        self.event_label = QLabel("Название мероприятия:")
        self.event_input = QLineEdit()
        self.event_input.setPlaceholderText("Введите название мероприятия")
        
        self.visitor_table = QTableWidget()
        self.visitor_table.setColumnCount(4)
        self.visitor_table.setHorizontalHeaderLabels(["ФИО", "Организация", "Время посещения", "Мероприятие"])
        
        self.export_button = QPushButton("Экспорт в Excel")
        self.export_button.setStyleSheet(
            "QPushButton {"
            "background-color: #2196F3;"
            "border: none;"
            "color: white;"
            "padding: 10px 24px;"
            "text-align: center;"
            "font-size: 16px;"
            "border-radius: 4px;"
            "}"
            "QPushButton:hover {"
            "background-color: #0b7dda;"
            "}"
        )
        
        right_layout.addWidget(self.info_label)
        right_layout.addWidget(self.event_label)
        right_layout.addWidget(self.event_input)
        right_layout.addWidget(self.visitor_table)
        right_layout.addWidget(self.export_button)
        
        # Инициализация базы данных
        self.init_db()
        
        # Настройка камеры
        self.camera = cv2.VideoCapture(0)
        self.timer = QTimer()
        self.scanning = False
        
        # Подключение сигналов
        self.scan_button.clicked.connect(self.toggle_scan)
        self.timer.timeout.connect(self.update_frame)
        self.export_button.clicked.connect(self.export_to_excel)
        
        # Добавляем виджет в стек
        self.stacked_widget.addWidget(self.scanner_widget)
    
    def init_generator_mode(self):
        """Инициализация режима генерации"""
        self.generator_widget = QWidget()
        generator_layout = QVBoxLayout(self.generator_widget)
        
        # 1. Секция загрузки файла
        file_frame = QFrame()
        file_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        file_layout = QHBoxLayout(file_frame)
        
        self.file_label = QLabel("Файл Excel не выбран")
        self.file_label.setStyleSheet("font-weight: bold; color: #3498db;")
        
        file_btn = QPushButton("Выбрать файл Excel")
        file_btn.setStyleSheet(
            "QPushButton {"
            "background-color: #3498db;"
            "}"
            "QPushButton:hover {"
            "background-color: #2980b9;"
            "}"
        )
        
        file_layout.addWidget(self.file_label)
        file_layout.addWidget(file_btn)
        generator_layout.addWidget(file_frame)
        
        # 2. Секция настроек
        settings_frame = QFrame()
        settings_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        settings_layout = QHBoxLayout(settings_frame)
        
        # Выбор листа
        sheet_layout = QVBoxLayout()
        sheet_label = QLabel("Лист:")
        self.sheet_combo = QComboBox()
        self.sheet_combo.setPlaceholderText("Выберите лист")
        sheet_layout.addWidget(sheet_label)
        sheet_layout.addWidget(self.sheet_combo)
        settings_layout.addLayout(sheet_layout)
        
        # Выбор колонки
        column_layout = QVBoxLayout()
        column_label = QLabel("Колонка с ФИО:")
        self.column_combo = QComboBox()
        self.column_combo.setPlaceholderText("Выберите колонку")
        column_layout.addWidget(column_label)
        column_layout.addWidget(self.column_combo)
        settings_layout.addLayout(column_layout)
        
        # Размер QR-кода
        size_layout = QVBoxLayout()
        size_label = QLabel("Размер QR:")
        self.size_spin = QSpinBox()
        self.size_spin.setRange(100, 500)
        self.size_spin.setValue(250)
        size_layout.addWidget(size_label)
        size_layout.addWidget(self.size_spin)
        settings_layout.addLayout(size_layout)
        
        generator_layout.addWidget(settings_frame)
        
        # 3. Секция генерации
        generate_btn = QPushButton("Сгенерировать QR-коды")
        generate_btn.setStyleSheet(
            "QPushButton {"
            "background-color: #e67e22;"
            "font-size: 16px;"
            "padding: 10px;"
            "}"
            "QPushButton:hover {"
            "background-color: #d35400;"
            "}"
        )
        generator_layout.addWidget(generate_btn, alignment=Qt.AlignCenter)
        
        # 4. Секция отображения данных
        data_frame = QFrame()
        data_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        data_layout = QHBoxLayout(data_frame)
        
        # Список ФИО
        self.data_list = QListWidget()
        self.data_list.itemClicked.connect(self.show_qr_code)
        data_layout.addWidget(self.data_list, 1)
        
        # Область просмотра QR-кода
        self.qr_display = QLabel()
        self.qr_display.setAlignment(Qt.AlignCenter)
        self.qr_display.setMinimumSize(350, 350)
        self.qr_display.setStyleSheet(
            "background-color: white; "
            "border: 2px solid #3498db;"
            "border-radius: 8px;"
            "padding: 10px;"
        )
        data_layout.addWidget(self.qr_display, 1)
        generator_layout.addWidget(data_frame)
        
        # 5. Секция сохранения
        save_frame = QFrame()
        save_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 15px;")
        save_layout = QHBoxLayout(save_frame)
        
        save_current_btn = QPushButton("Сохранить текущий QR-код")
        save_current_btn.setStyleSheet(
            "QPushButton {"
            "background-color: #3498db;"
            "}"
            "QPushButton:hover {"
            "background-color: #2980b9;"
            "}"
        )
        
        save_all_btn = QPushButton("Сохранить все QR-коды")
        save_all_btn.setStyleSheet(
            "QPushButton {"
            "background-color: #2ecc71;"
            "}"
            "QPushButton:hover {"
            "background-color: #27ae60;"
            "}"
        )
        
        save_layout.addWidget(save_current_btn)
        save_layout.addWidget(save_all_btn)
        generator_layout.addWidget(save_frame)
        
        # Основные переменные
        self.excel_file = None
        self.current_qr = None
        self.qr_codes = {}
        
        # Подключение сигналов
        file_btn.clicked.connect(self.load_excel_file)
        self.sheet_combo.currentTextChanged.connect(self.update_columns)
        generate_btn.clicked.connect(self.generate_qr_codes)
        save_current_btn.clicked.connect(self.save_current_qr)
        save_all_btn.clicked.connect(self.save_all_qr)
        
        # Добавляем виджет в стек
        self.stacked_widget.addWidget(self.generator_widget)
    
    # Методы для режима сканирования
    def init_db(self):
        self.conn = sqlite3.connect('visitors.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS visitors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                full_name TEXT,
                organization TEXT,
                qr_data TEXT UNIQUE,
                visit_time DATETIME,
                event TEXT
            )
        ''')
        self.conn.commit()
        self.update_visitors_table()
    
    def toggle_scan(self):
        if not self.scanning:
            self.scanning = True
            self.scan_button.setText("Остановить сканирование")
            self.scan_button.setStyleSheet(
                "QPushButton {"
                "background-color: #f44336;"
                "border: none;"
                "color: white;"
                "padding: 10px 24px;"
                "text-align: center;"
                "font-size: 16px;"
                "border-radius: 4px;"
                "}"
                "QPushButton:hover {"
                "background-color: #d32f2f;"
                "}"
            )
            self.timer.start(20)
        else:
            self.scanning = False
            self.scan_button.setText("Начать сканирование")
            self.scan_button.setStyleSheet(
                "QPushButton {"
                "background-color: #4CAF50;"
                "border: none;"
                "color: white;"
                "padding: 10px 24px;"
                "text-align: center;"
                "font-size: 16px;"
                "border-radius: 4px;"
                "}"
                "QPushButton:hover {"
                "background-color: #45a049;"
                "}"
            )
            self.timer.stop()
            self.camera_label.clear()
            self.camera_label.setText("Камера отключена")
    
    def update_frame(self):
        ret, frame = self.camera.read()
        if ret:
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            decoded_objects = decode(frame)
            
            if decoded_objects:
                qr_data = decoded_objects[0].data.decode('utf-8')
                self.process_qr_code(qr_data)
                self.timer.stop()
                self.scanning = False
                self.scan_button.setText("Начать сканирование")
                self.scan_button.setStyleSheet(
                    "QPushButton {"
                    "background-color: #4CAF50;"
                    "border: none;"
                    "color: white;"
                    "padding: 10px 24px;"
                    "text-align: center;"
                    "font-size: 16px;"
                    "border-radius: 4px;"
                    "}"
                    "QPushButton:hover {"
                    "background-color: #45a049;"
                    "}"
                )
            
            h, w, ch = frame_rgb.shape
            bytes_per_line = ch * w
            q_img = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
            self.camera_label.setPixmap(QPixmap.fromImage(q_img).scaled(
                self.camera_label.width(), 
                self.camera_label.height(), 
                Qt.KeepAspectRatio
            ))
    
    def process_qr_code(self, qr_data):
        try:
            self.cursor.execute("SELECT * FROM visitors WHERE qr_data=?", (qr_data,))
            visitor = self.cursor.fetchone()
            
            if visitor:
                visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                event = self.event_input.text() or "Не указано"
                self.cursor.execute(
                    "UPDATE visitors SET visit_time=?, event=? WHERE qr_data=?",
                    (visit_time, event, qr_data)
                )
                self.conn.commit()
                
                QMessageBox.information(
                    self, "Успешно", 
                    f"Посетитель {visitor[1]} уже зарегистрирован.\nВремя обновлено."
                )
            else:
                visit_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                event = self.event_input.text() or "Не указано"
                
                parts = qr_data.split(';')
                full_name = parts[0] if len(parts) > 0 else "Неизвестный"
                organization = parts[1] if len(parts) > 1 else "Не указана"
                
                self.cursor.execute(
                    "INSERT INTO visitors (full_name, organization, qr_data, visit_time, event) "
                    "VALUES (?, ?, ?, ?, ?)",
                    (full_name, organization, qr_data, visit_time, event)
                )
                self.conn.commit()
                
                QMessageBox.information(
                    self, "Успешно", 
                    f"Новый посетитель {full_name} зарегистрирован!"
                )
            
            self.update_visitors_table()
            
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
    
    def update_visitors_table(self):
        self.cursor.execute("SELECT full_name, organization, visit_time, event FROM visitors ORDER BY visit_time DESC LIMIT 50")
        visitors = self.cursor.fetchall()
        
        self.visitor_table.setRowCount(len(visitors))
        for row, visitor in enumerate(visitors):
            for col, data in enumerate(visitor):
                item = QTableWidgetItem(str(data))
                item.setTextAlignment(Qt.AlignCenter)
                self.visitor_table.setItem(row, col, item)
        
        self.visitor_table.resizeColumnsToContents()
    
    def export_to_excel(self):
        try:
            self.cursor.execute("SELECT full_name, organization, visit_time, event FROM visitors")
            visitors = self.cursor.fetchall()
            
            df = pd.DataFrame(visitors, columns=["ФИО", "Организация", "Время посещения", "Мероприятие"])
            df.to_excel("visitors.xlsx", index=False)
            
            QMessageBox.information(self, "Успешно", "Данные экспортированы в visitors.xlsx")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")
    
    # Методы для режима генерации
    def load_excel_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите файл Excel", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            self.excel_file = file_path
            self.file_label.setText(os.path.basename(file_path))
            
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить файл: {str(e)}")
    
    def update_columns(self):
        self.column_combo.clear()
        
        if not self.excel_file or not self.sheet_combo.currentText():
            return
            
        sheet_name = self.sheet_combo.currentText()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            for col_idx, cell in enumerate(sheet[1], 1):
                if cell.value:
                    self.column_combo.addItem(str(cell.value), col_idx)
                else:
                    self.column_combo.addItem(f"Колонка {col_idx}", col_idx)
            
            wb.close()
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось прочитать колонки: {str(e)}")
    
    def generate_qr_codes(self):
        if not all([self.excel_file, self.sheet_combo.currentText(), self.column_combo.currentText()]):
            QMessageBox.warning(self, "Ошибка", "Пожалуйста, выберите файл, лист и колонку")
            return
            
        self.data_list.clear()
        self.qr_codes = {}
        sheet_name = self.sheet_combo.currentText()
        column = self.column_combo.currentText()
        col_idx = self.column_combo.currentData()
        qr_size = self.size_spin.value()
        
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            sheet = wb[sheet_name]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= col_idx:
                    value = row[col_idx-1]
                    if value is not None:
                        value = str(value).strip()
                        if value:
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_H,
                                box_size=12,
                                border=4,
                            )
                            qr.add_data(value)
                            qr.make(fit=True)
                            
                            img = qr.make_image(fill_color="black", back_color="white")
                            img = img.resize((qr_size, qr_size), Image.Resampling.LANCZOS)
                            
                            self.qr_codes[value] = img
                            self.data_list.addItem(value)
            
            wb.close()
            QMessageBox.information(self, "Готово", f"Сгенерировано {len(self.qr_codes)} QR-кодов")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сгенерировать QR-коды: {str(e)}")
    
    def show_qr_code(self, item):
        text = item.text()
        if text in self.qr_codes:
            self.current_qr = self.qr_codes[text]
            
            img_byte_arr = io.BytesIO()
            self.current_qr.save(img_byte_arr, format='PNG')
            qimage = QImage.fromData(img_byte_arr.getvalue())
            pixmap = QPixmap.fromImage(qimage)
            
            display_size = min(self.qr_display.width(), self.qr_display.height()) - 20
            pixmap = pixmap.scaled(
                display_size, 
                display_size,
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            
            self.qr_display.setPixmap(pixmap)
    
    def save_current_qr(self):
        if not hasattr(self, 'current_qr') or not self.current_qr:
            QMessageBox.warning(self, "Ошибка", "Не выбран QR-код для сохранения")
            return
            
        current_item = self.data_list.currentItem()
        if not current_item:
            return
            
        text = current_item.text()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить QR-код", f"{text}.png", "PNG Files (*.png)"
        )
        
        if file_path:
            try:
                self.current_qr.save(file_path)
                QMessageBox.information(self, "Успешно", "QR-код сохранен")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл: {str(e)}")
    
    def save_all_qr(self):
        if not self.qr_codes:
            QMessageBox.warning(self, "Ошибка", "Нет QR-кодов для сохранения")
            return
            
        dir_path = QFileDialog.getExistingDirectory(self, "Выберите папку для сохранения")
        
        if dir_path:
            try:
                saved_count = 0
                for text, img in self.qr_codes.items():
                    safe_text = "".join(c for c in text if c.isalnum() or c in (' ', '_')).rstrip()
                    file_path = os.path.join(dir_path, f"{safe_text}.png")
                    img.save(file_path, "PNG")
                    saved_count += 1
                
                QMessageBox.information(self, "Успешно", f"Сохранено {saved_count} QR-кодов")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файлы: {str(e)}")
    
    def closeEvent(self, event):
        if hasattr(self, 'camera') and self.camera.isOpened():
            self.camera.release()
        if hasattr(self, 'conn'):
            self.conn.close()
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainApp()
    window.show()
    sys.exit(app.exec_())